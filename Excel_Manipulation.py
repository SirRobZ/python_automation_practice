from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font

"""
wb = Workbook()

ws = wb.active

ws['A1'] = "Hello World!"

wb.save('example.xlsx')


wb = load_workbook('example.xlsx')
ws = wb['Sheet']
print(ws['A1'].value)
 
"""

wb = load_workbook('example.xlsx')
ws = wb['Sheet']

"""
for cell in ws['A']:
    print(cell.value)


ws['B2'] = 'New Data'
wb.save('example.xlsx')


ws['A1'].font = Font(bold=True)
ws['C1'] = '=SUM(A1:B1)'

wb.save('example.xlsx')

"""
ws.merge_cells('A1:B1')
ws.freeze_panes = 'A2'

wb.save('example.xlsx')
